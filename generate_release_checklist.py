import os
import openpyxl
import argparse
from openpyxl.styles import Font, Alignment, Border, PatternFill

def generate_release_checklist(template_file_path, case_numbers, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    # Load the template workbook once
    template_wb = openpyxl.load_workbook(template_file_path)

    # Create the files for each case number
    for case_number in case_numbers:
        # Create a copy of the workbook
        new_wb = openpyxl.Workbook()
        for sheet_name in template_wb.sheetnames:
            template_ws = template_wb[sheet_name]
            new_ws = new_wb.create_sheet(sheet_name)
            
            # Copy the column widths
            for col in template_ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = template_ws.column_dimensions[col_letter].width
                new_ws.column_dimensions[col_letter].width = adjusted_width
            
            # Copy the content and formatting from the template worksheet to the new one
            for row in template_ws.iter_rows():
                for cell in row:
                    new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                    
                    # Copy font
                    if cell.font:
                        new_cell.font = Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            vertAlign=cell.font.vertAlign,
                            underline=cell.font.underline,
                            strike=cell.font.strike,
                            color=cell.font.color
                        )
                    
                    # Copy alignment
                    if cell.alignment:
                        new_cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            text_rotation=cell.alignment.text_rotation,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent
                        )
                    
                    # Copy border
                    if cell.border:
                        new_cell.border = Border(
                            left=cell.border.left,
                            right=cell.border.right,
                            top=cell.border.top,
                            bottom=cell.border.bottom
                        )
                    
                    # Copy fill (background color)
                    if cell.fill:
                        new_cell.fill = PatternFill(
                            fill_type=cell.fill.fill_type,
                            start_color=cell.fill.start_color,
                            end_color=cell.fill.end_color
                        )

        # Remove the default "Sheet" created during workbook creation
        if "Sheet" in new_wb.sheetnames:
            del new_wb["Sheet"]
        
        # Save the new workbook
        new_file_path = os.path.join(output_dir, f"{case_number}.xlsx")
        new_wb.save(new_file_path)
        print(f"Created file: {new_file_path}")

    print(f"All Excel files are created inside the '{output_dir}' folder")

    # Print summary to console
    print("\nThe code review checklist has been completed. Please review the following cases and let me know if anything is missing or requires further attention.")
    print("\n🚀Release:", output_dir.replace("_", " "))
    print("Total Cases:", len(case_numbers))
    print("\nIncluded Cases:")
    for case in case_numbers:
        print(case)

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--template_file_path", required=True, help="Path to the template file")
    parser.add_argument("--case_numbers", required=True, help="Case numbers for the checklist, separated by commas")
    parser.add_argument("--output_dir", required=True, help="Output directory where the Excel files will be saved")

    args = parser.parse_args()
    generate_release_checklist(template_file_path=args.template_file_path,
                               case_numbers=args.case_numbers.split(","),
                               output_dir=args.output_dir)